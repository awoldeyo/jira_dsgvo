from optparse import OptionParser

import pandas as pd
from pandas import ExcelWriter
from jira.exceptions import JIRAError
from jira.client import JIRA
from openpyxl import load_workbook
from tkinter import filedialog

from cocoa import Connection
from upload_resources import (excelname_mapping, dc_cols, daml_cols, 
                              dtype, heatmap, heatmap_area, template_cols, 
                              allNA_DC_cols)


class UploadIssues(object):
    '''Upload issues (DAML & DC) from Excel upload document.
       
       Parameters
       ----------
       filename : path to Excel upload document
    '''

    def __init__(self, filename):
        self.daml_blueprint = pd.read_pickle('./project_blueprints/daml_blueprint.pickle')
        self.dc_blueprint = pd.read_pickle('./project_blueprints/dc_blueprint.pickle')
        self.mandatory_DAML_cols = [c for c in daml_cols if "*" in c]
        self.mandatory_DC_cols = [c for c in dc_cols if "*" in c]
        self.filename = filename
        self.df = pd.read_excel(self.filename, skiprows=1, dtype=dtype)
        self.prepareFile()
    
    def prepareFile(self):
        '''Method to preprocess and create certain certain columns in Excel 
           upload document. This avoids errors when uploading issues to JIRA.
        '''
        
        # Format date to string
        self.df['Due-Date implemented'] = self.df['Due-Date implemented'].map(
                lambda x: x.strftime('%Y-%m-%d') if pd.notna(x) else ''
                )
        
        # Fill N/A with empty string for optional columns
        optional_cols = ['Sub Department', 
                         'Contact person (business department)', 
                         'Contact person (IT)', 
                         'Market'
                         ]
        self.df.loc[:, optional_cols].fillna('', inplace=True)
        
        # Cleanup data (remove trailing whitespace ...)
        self.df = self.df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
        
        # Create Heat-Map and Areas of activity Heatmap columns
        self.df['Heat-Map'] = self.df['Detailed Type*'].map(
                lambda x: heatmap.get(x, None)
                )
        self.df['Heat-Map'] = self.df['Heat-Map'].map(
                lambda x: f'{int(x)}' if pd.notna(x) else None
                )
        self.df['Areas of activity Heatmap'] = self.df['Detailed Type*'].map(
                lambda x: heatmap_area.get(x, None)
                )
        
        self.df['Linked Issue DC'] = None

        
    def createUploadDictDAML(self):
        '''Method to create dictionaries from Excel entries. Dictonaries can
           later be uploaded to DAML project. Entries for which mandatory 
           columns are not filled are considered as incomplete and will be 
           exported to Incomplete_data.xlsx.
        '''
        
        # Determine which DAML rows are valid 
        # (i.e. all mandatory columns are filled out)
        self.valid_DAML_rows = self.df.dropna(
                axis='index', 
                subset=self.mandatory_DAML_cols
                ).index
        
        # Determine DAML that can be uploaded and create dictionary for upload 
        condition_1 = (self.df.index.isin(self.valid_DAML_rows)) # 1. Mandatory columns are all filled
        condition_2 = (self.df['Linked Issue'].isna()) # 2. Issues have not been linked yet
        daml_filter = condition_1 & condition_2
        
        # Create upload dictionary
        self.df.loc[daml_filter, 'DAML dict'] = self.df[daml_filter].reindex(
                daml_cols, 
                axis=1).fillna('').apply(
                        lambda x: df_to_issuedict(x, self.daml_blueprint, 'DAML'), 
                        axis=1)
        
        # Determine incomplete rows
        incomplete_rows = self.df.loc[daml_filter==False].index
        
        # Create dataframe from incomplete row(s)
        self.incomplete_df = self.df.loc[incomplete_rows, :]
        
        # Drop incomplete row(s) from original dataframe
        self.df.drop(labels=incomplete_rows, axis=0, inplace=True)
        
        # Store incorrect dataframe in new Upload template
        book = load_workbook('template/UpoadFile2_version_1.3.xlsm')
        writer = ExcelWriter('output/Incomplete_data.xlsx', engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        self.incomplete_df.to_excel(
                excel_writer=writer,
                sheet_name='Upload',
                header=False,
                index=False,
                startrow=2
                )
        writer.save()

    def createUploadDictDC(self):
        '''Method to create dictionaries from Excel entries. Dictonaries can
           later be uploaded to DC project. Entries for which mandatory 
           columns are not filled are considered as incomplete and will be 
           exported to Incomplete_data.xlsx.
        '''
        self.DC_notAllNA = self.df.dropna(axis=0, 
                                          how='all', 
                                          subset=allNA_DC_cols).index
         # Determine which DC rows are valid 
         # (i.e. all mandatory columns are filled out)
        self.valid_DC_rows = self.df.loc[self.DC_notAllNA, :].dropna(
                axis='index',
                how='any',
                subset=self.mandatory_DC_cols
                ).index
                 
        # Determine DC that can be uploaded and create dictionary for upload 
        condition_1 = (self.df.index.isin(self.valid_DC_rows)) # 1. Mandatory columns are all filled
        condition_2 = (self.df['Linked Issue DC'].isna()) # 2. Issues have not been linked yet
        dc_filter = condition_1 & condition_2
         
         # Create upload dictionary
        self.df.loc[dc_filter, 'DC dict'] = self.df[dc_filter].reindex(
                dc_cols, 
                axis=1).fillna('').apply(
                        lambda x: df_to_issuedict(x, self.dc_blueprint, 'DC'),
                        axis=1)
         
         # Determine incomplete rows
        condition_3 = self.df['Linked Issue'].notna() # Linked Issue, i.e. corresponding DAML issue exists
        condition_4 = (dc_filter == False) # Not in dc_filter (i.e. mandatory field not filled and not linked yet)
        condition_5 = self.df.index.isin(self.DC_notAllNA) # DCs are not all na
        incomplete_rows = self.df.loc[condition_3 & condition_4 & condition_5].index
         
         # Add incomplete row(s) to existing incomplete dataframe
        self.incomplete_df = pd.concat(
                (self.incomplete_df, self.df.loc[incomplete_rows, :]), 
                sort=True
                )
         
        # Drop incomplete row(s) from original dataframe
        self.df.drop(labels=incomplete_rows, axis=0, inplace=True)
         
         # Reindex incomplete dataframe columns to match original template
        self.incomplete_df = self.incomplete_df.reindex(
                labels=template_cols,
                axis=1
                )
         
        book = load_workbook('template/UpoadFile2_version_1.3.xlsm')
        writer = ExcelWriter('output/Incomplete_data.xlsx', engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        self.incomplete_df.to_excel(
                excel_writer=writer,
                sheet_name='Upload',
                header=False,
                index=False,
                startrow=2
                )
        writer.save()


    def postDAML(self):
        '''Method to upload DAML dictionaries as issues to JIRA. Issues not 
           successfully uploaded will be exported to Incorrect_data.xlsx.
        '''
        self.df['results'] = self.df['DAML dict'].map(post_issues)
        results = self.df.apply(lambda x: x['results'], 
                                axis=1, 
                                result_type='expand').copy()
        results.columns = ['Linked Issue', 
                           'DAML_upload_success', 
                           'DAML_Error_message'
                           ]
        self.df.drop(labels=['Linked Issue'], 
                     axis=1, 
                     inplace=True)
        self.df = pd.concat((self.df, results), 
                            axis=1).drop(
                                    labels=['results'], axis=1
                                    )
        
        # Create incorrect dataframe for not successfully uploaded DAML issues
        self.incorrect_df = self.df.loc[self.df.DAML_upload_success==False,:].copy(True)
        
        # Drop incorrect rows from original dataframe
        self.df.drop(
                labels=self.incorrect_df.index, 
                axis=0, 
                inplace=True)
        
        # Drop irrelevant columns from incorrect dataframe
        self.incorrect_df.drop(labels=['Heat-Map',
                                       'Linked Issue DC',
                                       'Areas of activity Heatmap',
                                       'DAML dict',
                                       'DAML_upload_success',
                                       ], 
                                axis=1, 
                                inplace=True
                                )
        
        # Store incorrect dataframe in new Upload template
        book = load_workbook('template/UpoadFile2_version_1.3.xlsm')
        writer = ExcelWriter('output/Incorrect_data.xlsx', engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        self.incorrect_df.to_excel(
                excel_writer=writer,
                sheet_name='Upload',
                header=False,
                index=False,
                startrow=2
                )
        writer.save()
    
    def postDC(self):
        '''Method to upload DC dictionaries as issues to JIRA. Issues not 
           successfully uploaded will be exported to Incorrect_data.xlsx.
        '''        
        DC_filter = self.df['DC dict'].notna()
        self.df.loc[DC_filter, 'results'] = self.df.loc[DC_filter, 'DC dict'].map(post_issues)
        results = self.df.loc[DC_filter,:].apply(lambda x: x['results'], 
                                axis=1, 
                                result_type='expand').copy()
        results.columns = ['Linked Issue DC', 
                           'DC_upload_success', 
                           'DC_Error_message'
                           ]
        
        self.df.drop(labels=['Linked Issue DC'], 
                     axis=1, 
                     inplace=True)
        
        self.df = pd.concat((self.df, results), 
                            axis=1).drop(
                                    labels=['results'], axis=1
                                    )
        
        # Add not successfully uploaded DC issues to incorrect dataframe
        self.incorrect_df_DC = self.df.loc[self.df.DC_upload_success==False,:]
        if self.incorrect_df_DC.shape[0] >=1:
            self.incorrect_df = pd.concat(
                    (self.incorrect_df, self.incorrect_df_DC),
                    sort=True,
                    )
        
            # Drop incorrect rows from original dataframe
            self.df.drop(
                    labels=self.incorrect_df_DC.index, 
                    axis=0, 
                    inplace=True)
            
            self.incorrect_df.drop(labels=['Heat-Map',
                                           'Areas of activity Heatmap',
                                           'DC dict',
                                           'DC_upload_success',
                                           ], 
                                    axis=1, 
                                    inplace=True
                                    )
        
        # Adjust incorrect dataframe column to template column
        incorrect_cols = template_cols
        self.incorrect_df = self.incorrect_df.reindex(labels=incorrect_cols,
                                                      axis=1,
                                                      )
        
        # Store incorrect dataframe in new Upload template
        book = load_workbook('template/UpoadFile2_version_1.3.xlsm')
        writer = ExcelWriter('output/Incorrect_data.xlsx', engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        self.incorrect_df.to_excel(
                excel_writer=writer,
                sheet_name='Upload',
                header=False,
                index=False,
                startrow=2
                )
        writer.save()

    def addCommentDAML(self):
        '''Method to add values in column 'Comment DAML' as comment
           to issues specified in column 'Linked Issue'.
        '''
        comment_filter = self.df['Comment DAML'].notna() & self.df['Linked Issue'].notna()
        _ = self.df.loc[comment_filter, :].apply(
                lambda x: jira.add_comment(issue=x['Linked Issue'], body=x['Comment DAML']), 
                axis=1)
    
    def addCommentDC(self):
        '''Method to add values in column 'Comment DC' as comment
           to issues specified in column 'Linked Issue DC'.
        '''
        comment_filter = self.df['Comment'].notna() & self.df['Linked Issue DC'].notna()
        _ = self.df.loc[comment_filter, :].apply(
        lambda x: jira.add_comment(issue=x['Linked Issue DC'], body=x['Comment']), 
        axis=1)
    
    def changeStatusDAML(self):
        '''Method to change DAML issues' status to value in column 'Status DAML'.
        '''
        status_filter = self.df['Status DAML'].notna() & self.df['Linked Issue'].notna()
        self.df.loc[status_filter, 'DAML object'] = self.df.loc[status_filter, 'Linked Issue'].map(lambda x: get_issues(x))
        self.df.loc[status_filter, 'Status DAML'] = self.df.loc[status_filter, 'Status DAML'].str.title()
        _ = self.df.loc[status_filter].apply(
                lambda x: change_daml_status(x), 
                axis=1)
    
    def changeStatusDC(self):
        '''Method to change DC issues' status to value in column 'Status'.
        '''
        status_filter = self.df['Status'].notna() & self.df['Linked Issue DC'].notna()
        self.df.loc[status_filter, 'DC object'] = self.df.loc[status_filter, 'Linked Issue DC'].map(lambda x: get_issues(x))
        self.df.loc[status_filter, 'Status'] = self.df.loc[status_filter, 'Status'].str.title()
        _ = self.df.loc[status_filter].apply(
                lambda x: change_dc_status(x), 
                axis=1)
    
    def linkDAML_DC(self):
        '''Method to link DC and DAML issues which both are listed in one row and
           were uploaded sucessfully.
        '''
        link_filter = self.df['Linked Issue'].notna() & self.df['Linked Issue DC'].notna()
        _ = self.df.loc[link_filter, :].apply(
                lambda x: jira.create_issue_link(type='Relates', inwardIssue=x['Linked Issue'], outwardIssue=x['Linked Issue DC']), 
                axis=1)

def from_blueprint(blueprint, field, fieldvalue):
    '''Looks up and returns required data format'''
    index = excelname_mapping.get(field, None)
    fieldtype = blueprint.loc[index,'schema']
    attribute_type = blueprint.loc[index,'attribute_type']
    
    if fieldtype == 'array':
        if pd.notna(attribute_type):
            return [{f'{attribute_type}': fieldvalue}]
        else:
            return fieldvalue.split(',')
    elif fieldtype in ['priority', 'option', 'issuetype']:
        return {f'{attribute_type}': fieldvalue}
    elif fieldtype in ['string', 'date']:
        return fieldvalue
    elif fieldtype == 'user':
        return {'name': fieldvalue}

def df_to_issuedict(item, blueprint, project):
    '''Returns issue dictionaries'''
    issue_dict = {excelname_mapping[c]:from_blueprint(blueprint=blueprint, field=c, fieldvalue=item[c]) for c in item.index}
    if project == 'DAML':
        issue_dict['project'] = {'id': '11605'}
        issue_dict['issuetype'] = {'name': 'Defect'}
        #issue_dict['labels'] = ['VW-PKW'] # uncomment after deployment
        issue_dict['labels'] = ['Testing'] # comment out after deployment
    elif project == 'DC':
        issue_dict['project'] = {'id': '12302'}
        issue_dict['issuetype'] = {'name': 'Task'}
    return issue_dict

def post_issues(x):
    '''Posts Jira issues and returns True/False and Object Key/Error text'''
    try:
        obj = jira.create_issue(x)
        obj_success, obj_key, obj_error= True, obj.key, None
    except JIRAError as j:
        obj_success, obj_key, obj_error = False, None, j.text
    finally:
        return obj_key, obj_success, obj_error
    
def get_issues(key):
    try:
        obj = jira.search_issues(jql_str=f'key={key}', maxResults=False)
        obj = obj[0]
        return obj
    except JIRAError as j:
        print(j)
        return None
    except KeyError as k:
        print(f'Could not find issue for {key}.')
        return None
    
def change_daml_status(x):
    if x['Status DAML'] == 'New':
        pass
    elif x['Status DAML'] == 'In Assessment':
        jira.transition_issue(x['DAML object'], transition='11') # 1. Approve
        
    elif x['Status DAML'] == 'In Progress':
        jira.transition_issue(x['DAML object'], transition='11') # 1. Aprove
        jira.transition_issue(x['DAML object'], transition='21') # 2. Implement
        
    elif x['Status DAML'] == 'Done':
        jira.transition_issue(x['DAML object'], transition='11') # 1. Aprove
        jira.transition_issue(x['DAML object'], transition='21') # 2. Implement
        jira.transition_issue(x['DAML object'], transition='41') # 3. Aprove Concept
        
    elif x['Status DAML'] == 'Rejected':
        jira.transition_issue(x['DAML object'], transition='11') # 1. Aprove
        jira.transition_issue(x['DAML object'], transition='21') # 2. Implement
        jira.transition_issue(x['DAML object'], transition='31') # 2. Implement
        
def change_dc_status(x):
    if x['Status'] == 'Draft':
        pass
    elif x['Status'] == 'Order Approval':
        jira.transition_issue(x['DC object'], transition='21') # 1. Approve
        
    elif x['Status'] == 'Implemented':
        jira.transition_issue(x['DC object'], transition='21') # 1. Aprove
        jira.transition_issue(x['DC object'], transition='11') # 2. Implement
        
    elif x['Status'] == 'Resolved':
        jira.transition_issue(x['DC object'], transition='21') # 1. Aprove
        jira.transition_issue(x['DC object'], transition='11') # 2. Implement
        jira.transition_issue(x['DC object'], transition='141') # 3. Aprove Concept
        jira.transition_issue(x['DC object'], transition='101') # 4. Resolve
        
    elif x['Status'] == 'Closed':
        jira.transition_issue(x['DC object'], transition='21') # 1. Aprove
        jira.transition_issue(x['DC object'], transition='11') # 2. Implement
        jira.transition_issue(x['DC object'], transition='141') # 3. Aprove Concept
        jira.transition_issue(x['DC object'], transition='101') # 4. Resolve
        jira.transition_issue(x['DC object'], transition='121') # 5. Close


# Authentication:
try:
    # First try to connect with existing cookie
    jira = Connection(stored_cookie=True).jira
    if not isinstance(jira, JIRA):
        raise ValueError
except (ValueError, FileNotFoundError):
    # If cookie expired request username and password
    jira = Connection().jira


filename = filedialog.askopenfile()

up = UploadIssues(filename.name)

up.createUploadDictDAML()
up.postDAML()
up.addCommentDAML() 
up.changeStatusDAML()


up.createUploadDictDC()
up.postDC()
up.addCommentDC() 
up.changeStatusDC()

up.linkDAML_DC()

# =============================================================================
# parser = OptionParser()
# args = parser.add_option("-c", action="store_true", dest="stored", help="Use stored cookie to authenticate.")
# (options, args) = parser.parse_args()
# jira = Connection(stored_cookie=options.stored).jira
# =============================================================================
